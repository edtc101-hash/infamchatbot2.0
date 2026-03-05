"""
b.ver1.xlsx 엑셀 데이터 → 챗봇 학습 데이터(learned-data.json) 변환
- Sheet1: 제품 지식, 시공 팁, 배송 규격
- 궁금: 고객 질문과 전문가 답변
- 생각: 고객 피드백, 설문 결과
"""
import openpyxl
import json
import os
import re
from datetime import datetime

EXCEL_PATH = r'c:\Users\kimdy\인팸 쳇봇\data\b.ver1.xlsx'
LEARNED_PATH = r'c:\Users\kimdy\인팸 쳇봇\learned-data.json'

def extract_keywords(text):
    """한국어 키워드 추출"""
    stop = ['은', '는', '이', '가', '을', '를', '에', '의', '로', '으로', '과', '와', '하다', '있다', '없다', '되다', '않다',
            '것', '수', '등', '및', '또는', '대한', '위해', '경우', '때문', '있습니다', '합니다', '드립니다']
    words = re.sub(r'[^\uAC00-\uD7A3a-zA-Z0-9\s]', ' ', text).split()
    words = [w for w in words if len(w) >= 2 and w not in stop]
    return list(dict.fromkeys(words))[:10]  # 중복 제거, 최대 10개

def categorize(text):
    """텍스트 내용 기반 카테고리 분류"""
    t = text.lower()
    if any(k in t for k in ['시공', '부착', '접착', '설치', '하지작업', '마감', '절곡']):
        return '시공 가이드'
    if any(k in t for k in ['배송', '용달', '택배', '배달', '차량', '적재']):
        return '배송/물류'
    if any(k in t for k in ['가격', '견적', '단가']):
        return '가격/견적'
    if any(k in t for k in ['소프트스톤', '소프트 스톤']):
        return '소프트 스톤'
    if any(k in t for k in ['카빙', 'cv']):
        return '카빙 스톤'
    if any(k in t for k in ['스텐', '플레이트']):
        return '스텐 플레이트'
    if any(k in t for k in ['wpc', 'spc', '월패널', '패널']):
        return 'WPC/SPC 월패널'
    if any(k in t for k in ['시멘트', '블럭', '블록']):
        return '시멘트 블럭'
    if any(k in t for k in ['방염', '인증', '성적서']):
        return '인증/방염'
    if any(k in t for k in ['외장', '외부']):
        return '외장재'
    if any(k in t for k in ['커스텀', '커스터마이징', '맞춤']):
        return '커스터마이징'
    if any(k in t for k in ['샘플', '견본']):
        return '샘플'
    if any(k in t for k in ['고객', '평가', '만족', '불만', '피드백']):
        return '고객 피드백'
    return '제품 지식'

def make_qa_from_statement(text):
    """단일 문장 → Q&A 형태로 변환"""
    t = text.strip()
    if not t or len(t) < 10:
        return None

    # 이미 Q&A 형태인 경우
    if '?' in t or '?' in t:
        parts = re.split(r'[?？]', t, 1)
        if len(parts) == 2 and len(parts[1].strip()) > 5:
            return {'question': parts[0].strip() + '?', 'answer': parts[1].strip()}

    # "=> " 또는 "= >" 구분자
    if '=>' in t or '=>' in t:
        parts = re.split(r'=>|=>', t, 1)
        if len(parts) == 2:
            return {'question': parts[0].strip(), 'answer': parts[1].strip()}

    # 단순 정보 → "~에 대해 알려주세요" 형태
    cat = categorize(t)
    keywords = extract_keywords(t)
    q_hint = keywords[0] if keywords else cat
    return {
        'question': f'{q_hint} 관련 정보',
        'answer': t
    }

print("Loading Excel...", flush=True)
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)

all_entries = []
entry_id_counter = 0

# ================================
# Sheet1: 제품 지식/시공/배송 정보
# ================================
ws1 = wb[wb.sheetnames[0]]
print(f"\n처리 중: {wb.sheetnames[0]}", flush=True)

buffer_lines = []
for row in ws1.iter_rows(values_only=True):
    cells = [str(c).strip() if c is not None else '' for c in row[:4]]
    text = ' '.join(c for c in cells if c).strip()
    
    if not text or len(text) < 8:
        # 빈 줄 → 버퍼 플러시
        if buffer_lines:
            combined = '\n'.join(buffer_lines)
            qa = make_qa_from_statement(combined)
            if qa and len(qa['answer']) >= 10:
                entry_id_counter += 1
                cat = categorize(combined)
                all_entries.append({
                    'id': f'excel_s1_{entry_id_counter}',
                    'category': cat,
                    'question': qa['question'],
                    'answer': qa['answer'],
                    'keywords': extract_keywords(combined),
                    'source': 'excel_b_ver1',
                    'sheet': wb.sheetnames[0],
                    'learnedAt': datetime.now().isoformat()
                })
            buffer_lines = []
    else:
        buffer_lines.append(text)

# 마지막 버퍼 플러시
if buffer_lines:
    combined = '\n'.join(buffer_lines)
    qa = make_qa_from_statement(combined)
    if qa and len(qa['answer']) >= 10:
        entry_id_counter += 1
        all_entries.append({
            'id': f'excel_s1_{entry_id_counter}',
            'category': categorize(combined),
            'question': qa['question'],
            'answer': qa['answer'],
            'keywords': extract_keywords(combined),
            'source': 'excel_b_ver1',
            'sheet': wb.sheetnames[0],
            'learnedAt': datetime.now().isoformat()
        })

print(f"  Sheet1: {entry_id_counter}개 항목 추출", flush=True)

# ================================
# 궁금 Sheet: 고객 Q&A
# ================================
if '궁금' in wb.sheetnames:
    ws_q = wb['궁금']
    print("\n처리 중: 궁금", flush=True)
    q_count = 0
    
    for row in ws_q.iter_rows(values_only=True):
        cells = [str(c).strip() if c is not None else '' for c in row[:2]]
        text = cells[0]
        extra = cells[1] if len(cells) > 1 else ''
        
        if not text or len(text) < 10:
            continue
        
        # Q&A 형태 또는 정보
        qa = make_qa_from_statement(text + (' ' + extra if extra else ''))
        if qa and len(qa['answer']) >= 10:
            entry_id_counter += 1
            q_count += 1
            all_entries.append({
                'id': f'excel_q_{entry_id_counter}',
                'category': categorize(text),
                'question': qa['question'],
                'answer': qa['answer'],
                'keywords': extract_keywords(text + ' ' + extra),
                'source': 'excel_b_ver1',
                'sheet': '궁금',
                'learnedAt': datetime.now().isoformat()
            })
    
    print(f"  궁금: {q_count}개 항목 추출", flush=True)

# ================================
# 생각 Sheet: 고객 피드백/설문
# ================================
if '생각' in wb.sheetnames:
    ws_i = wb['생각']
    print("\n처리 중: 생각", flush=True)
    i_count = 0
    
    for row in ws_i.iter_rows(values_only=True):
        cells = [str(c).strip() if c is not None else '' for c in row[:4]]
        text = ' '.join(c for c in cells if c).strip()
        
        if not text or len(text) < 15:
            continue
        
        # 설문/피드백 관련 내용만 추출
        if any(k in text for k in ['고객', '만족', '불만', '시공', '자재', '설문', '니즈', '제품', '평가', '피드백']):
            qa = make_qa_from_statement(text)
            if qa and len(qa['answer']) >= 10:
                entry_id_counter += 1
                i_count += 1
                all_entries.append({
                    'id': f'excel_i_{entry_id_counter}',
                    'category': '고객 피드백',
                    'question': qa['question'],
                    'answer': qa['answer'],
                    'keywords': extract_keywords(text),
                    'source': 'excel_b_ver1',
                    'sheet': '생각',
                    'learnedAt': datetime.now().isoformat()
                })
    
    print(f"  생각: {i_count}개 항목 추출", flush=True)

wb.close()

# ================================
# 기존 learned-data.json에 병합
# ================================
print(f"\n총 {len(all_entries)}개 새 항목 추출 완료", flush=True)

# 기존 데이터 로드
existing = []
if os.path.exists(LEARNED_PATH):
    with open(LEARNED_PATH, 'r', encoding='utf-8') as f:
        existing = json.load(f)

# 기존 엑셀 데이터 제거 (재실행 시 중복 방지)
existing = [d for d in existing if d.get('source') != 'excel_b_ver1']
print(f"기존 학습 데이터: {len(existing)}개 (엑셀 외)", flush=True)

# 병합
merged = existing + all_entries
with open(LEARNED_PATH, 'w', encoding='utf-8') as f:
    json.dump(merged, f, ensure_ascii=False, indent=2)

print(f"\n✅ learned-data.json 저장 완료: 총 {len(merged)}개 항목", flush=True)
print(f"   - 기존: {len(existing)}개")
print(f"   - 엑셀 신규: {len(all_entries)}개")

# 카테고리별 통계
cats = {}
for e in all_entries:
    c = e['category']
    cats[c] = cats.get(c, 0) + 1
print("\n카테고리별 분포:")
for c, n in sorted(cats.items(), key=lambda x: -x[1]):
    print(f"  {c}: {n}개")
